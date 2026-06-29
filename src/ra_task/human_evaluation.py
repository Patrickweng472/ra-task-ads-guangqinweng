from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path
from typing import Literal

import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.datavalidation import DataValidation

from .analysis import quadratic_weighted_kappa
from .llm_labeling import content_text, rule_score

RUBRIC_VERSION = "2.1.0"
DEFAULT_REVIEW_SEED = 20260629
REVIEW_COLUMNS = [
    "review_id",
    "岗位",
    "岗位描述",
    "岗位标签",
    "human_score",
    "technology_role",
    "strict_ai",
    "human_confidence",
    "human_evidence",
    "human_note",
]
BLIND_SOURCE_COLUMNS = ["review_id", "岗位", "岗位描述", "岗位标签"]
TECHNOLOGY_ROLES = {"none", "auxiliary", "core"}
CONFIDENCE_VALUES = {"high", "medium", "low"}
AI_CANDIDATE_RE = re.compile(
    r"人工智能|机器学习|深度学习|神经网络|计算机视觉|机器视觉|模式识别|自然语言|\bNLP\b|大模型|\bLLM\b|生成式\s*AI|推荐系统|算法|ISP",
    re.IGNORECASE,
)


def score_from_dimensions(technology_role: str, strict_ai: bool) -> int:
    """Apply the frozen v2.1 deterministic construct-to-score mapping."""
    if technology_role not in TECHNOLOGY_ROLES:
        raise ValueError(f"invalid technology_role: {technology_role}")
    if strict_ai and technology_role != "core":
        raise ValueError("strict_ai=true requires technology_role=core")
    if technology_role == "none":
        return 0
    if technology_role == "auxiliary":
        return 1
    return 3 if strict_ai else 2


def _review_id(split: str, canonical_id: str, seed: int) -> str:
    digest = hashlib.sha256(f"{RUBRIC_VERSION}|{seed}|{split}|{canonical_id}".encode()).hexdigest()[:10].upper()
    return f"{'DEV' if split == 'development' else 'HOLD'}-{digest}"


def _bool_column(series: pd.Series) -> pd.Series:
    return series.astype(str).str.casefold().eq("true")


def build_sampling_frame(
    ads: pd.DataFrame,
    labels: pd.DataFrame,
    version_comparison: pd.DataFrame,
    reliability_sample: pd.DataFrame,
) -> pd.DataFrame:
    """Build private sampling features. None of these features enter blind sheets."""
    frame = ads[["canonical_id", "岗位", "岗位描述", "岗位标签"]].copy()
    frame["canonical_id"] = frame["canonical_id"].astype(str)
    current = labels[["canonical_id", "score", "confidence"]].copy()
    current["canonical_id"] = current["canonical_id"].astype(str)
    current["score"] = current["score"].astype(int)
    frame = frame.merge(current, on="canonical_id", validate="one_to_one")

    comparison = version_comparison[["canonical_id", "score_changed", "crossed_main_threshold"]].copy()
    comparison["canonical_id"] = comparison["canonical_id"].astype(str)
    comparison["score_changed"] = _bool_column(comparison["score_changed"])
    comparison["crossed_main_threshold"] = _bool_column(comparison["crossed_main_threshold"])
    frame = frame.merge(comparison, on="canonical_id", how="left", validate="one_to_one")

    audit = reliability_sample[["canonical_id", "primary_score", "audit_score", "selection_reason"]].copy()
    audit["canonical_id"] = audit["canonical_id"].astype(str)
    audit["audit_disagreement"] = audit["primary_score"].astype(int).ne(audit["audit_score"].astype(int))
    audit = audit[["canonical_id", "audit_disagreement", "selection_reason"]]
    frame = frame.merge(audit, on="canonical_id", how="left", validate="one_to_one")
    frame["audit_disagreement"] = frame["audit_disagreement"].fillna(False).astype(bool)
    frame["selection_reason"] = frame["selection_reason"].fillna("")
    frame["score_changed"] = frame["score_changed"].fillna(False).astype(bool)
    frame["crossed_main_threshold"] = frame["crossed_main_threshold"].fillna(False).astype(bool)

    frame["source_text"] = frame.apply(content_text, axis=1)
    frame["rule_score"] = frame["source_text"].map(lambda text: rule_score(text)[0])
    frame["rule_model_threshold_conflict"] = frame["rule_score"].ge(2).ne(frame["score"].ge(2))
    frame["low_confidence"] = frame["confidence"].eq("low")
    frame["ai_algorithm_candidate"] = frame["source_text"].str.contains(AI_CANDIDATE_RE)
    frame["risk_points"] = (
        frame["audit_disagreement"].astype(int) * 8
        + frame["crossed_main_threshold"].astype(int) * 7
        + frame["rule_model_threshold_conflict"].astype(int) * 6
        + frame["low_confidence"].astype(int) * 5
        + frame["score_changed"].astype(int) * 3
        + frame["ai_algorithm_candidate"].astype(int) * 2
        + frame["score"].isin([1, 2]).astype(int)
    )
    return frame


def _randomized(frame: pd.DataFrame, seed: int) -> pd.DataFrame:
    return frame.sample(frac=1, random_state=seed).reset_index(drop=True)


def _take_holdout(frame: pd.DataFrame, excluded_ids: set[str], seed: int) -> pd.DataFrame:
    quotas = {0: 18, 1: 20, 2: 21, 3: 1}
    available = frame.loc[~frame["canonical_id"].isin(excluded_ids)].copy()
    selected: list[pd.DataFrame] = []
    for score, quota in quotas.items():
        pool = available.loc[available["score"].eq(score)].copy()
        candidates = _randomized(pool.loc[pool["ai_algorithm_candidate"]], seed + score * 31 + 1)
        non_candidates = _randomized(pool.loc[~pool["ai_algorithm_candidate"]], seed + score * 31 + 2)
        candidate_target = min(len(candidates), max(1, quota // 2))
        first = candidates.head(candidate_target)
        second = non_candidates.head(quota - len(first))
        chosen = pd.concat([first, second], ignore_index=True)
        if len(chosen) < quota:
            chosen_ids = set(chosen["canonical_id"])
            fill = _randomized(pool.loc[~pool["canonical_id"].isin(chosen_ids)], seed + score * 31 + 3)
            chosen = pd.concat([chosen, fill.head(quota - len(chosen))], ignore_index=True)
        if len(chosen) != quota:
            raise ValueError(f"cannot fill holdout quota for score {score}: {len(chosen)}/{quota}")
        selected.append(chosen)
    return pd.concat(selected, ignore_index=True)


def _take_development(frame: pd.DataFrame, seed: int) -> pd.DataFrame:
    quotas = {0: 15, 1: 20, 2: 24, 3: 1}
    available = frame.copy()
    available["tie_break"] = available["canonical_id"].map(
        lambda item_id: hashlib.sha256(f"{seed}|development|{item_id}".encode()).hexdigest()
    )
    selected: list[pd.DataFrame] = []
    for score, quota in quotas.items():
        pool = available.loc[available["score"].eq(score)].sort_values(
            ["risk_points", "tie_break"], ascending=[False, True]
        )
        candidate_reserve = {0: 1, 1: 1, 2: 4, 3: 1}[score]
        boundary_cases = pool.loc[pool["ai_algorithm_candidate"]].head(candidate_reserve)
        remainder = pool.loc[~pool["canonical_id"].isin(boundary_cases["canonical_id"])]
        chosen = pd.concat([boundary_cases, remainder.head(quota - len(boundary_cases))], ignore_index=True)
        if len(chosen) != quota:
            raise ValueError(f"cannot fill development quota for score {score}: {len(chosen)}/{quota}")
        selected.append(chosen)
    return pd.concat(selected, ignore_index=True)


def select_human_reference_samples(features: pd.DataFrame, seed: int = DEFAULT_REVIEW_SEED) -> pd.DataFrame:
    """Select a risk-enriched development set, then stratify the locked remainder."""
    development = _take_development(features, seed).assign(split="development")
    holdout = _take_holdout(features, set(development["canonical_id"]), seed).assign(split="holdout")
    selected = pd.concat([development, holdout], ignore_index=True)
    if len(selected) != 120 or selected["canonical_id"].duplicated().any():
        raise ValueError("human reference selection must contain 120 disjoint ads")
    selected["review_id"] = selected.apply(
        lambda row: _review_id(str(row["split"]), str(row["canonical_id"]), seed), axis=1
    )
    if selected["review_id"].duplicated().any():
        raise ValueError("review_id collision")
    return selected


def blind_review_frame(selected: pd.DataFrame, split: Literal["development", "holdout"], seed: int) -> pd.DataFrame:
    subset = selected.loc[selected["split"].eq(split), ["review_id", "岗位", "岗位描述", "岗位标签"]].copy()
    subset = _randomized(subset, seed + (11 if split == "development" else 29))
    for column in REVIEW_COLUMNS[4:]:
        subset[column] = ""
    return subset[REVIEW_COLUMNS]


def _safe_text_cell(cell: object, value: object) -> None:
    cell.value = "" if value is None else str(value)
    cell.data_type = "s"


def write_review_workbook(frame: pd.DataFrame, path: Path, *, split: Literal["development", "holdout"]) -> None:
    """Write a blind, validated, professionally formatted reviewer workbook."""
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "待审核"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "E2"
    sheet.auto_filter.ref = f"A1:J{len(frame) + 1}"
    sheet.row_dimensions[1].height = 30

    navy = "17365D"
    light_blue = "D9EAF7"
    pale_yellow = "FFF2CC"
    pale_red = "FCE4D6"
    white = "FFFFFF"
    thin = Side(style="thin", color="D9E2F3")
    input_columns = set(REVIEW_COLUMNS[4:])
    for column_index, column in enumerate(REVIEW_COLUMNS, start=1):
        cell = sheet.cell(1, column_index, column)
        cell.font = Font(name="等线", size=11, bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=navy))
        cell.comment = Comment(
            "人工填写字段" if column in input_columns else "盲审材料：请勿尝试识别公司、年份或模型结果",
            "RA Task",
        )

    for row_index, row in enumerate(frame.itertuples(index=False, name=None), start=2):
        sheet.row_dimensions[row_index].height = 72
        for column_index, value in enumerate(row, start=1):
            cell = sheet.cell(row_index, column_index)
            _safe_text_cell(cell, value)
            cell.font = Font(name="等线", size=10, color="1F1F1F")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
            if REVIEW_COLUMNS[column_index - 1] in input_columns:
                cell.fill = PatternFill("solid", fgColor=pale_yellow)
                cell.protection = Protection(locked=False)
            elif row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F7FAFC")

    widths = {"A": 18, "B": 24, "C": 82, "D": 35, "E": 13, "F": 18, "G": 12, "H": 18, "I": 42, "J": 42}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    score_validation = DataValidation(type="whole", operator="between", formula1="0", formula2="3", allow_blank=True)
    role_validation = DataValidation(type="list", formula1='"none,auxiliary,core"', allow_blank=True)
    ai_validation = DataValidation(type="list", formula1='"false,true"', allow_blank=True)
    confidence_validation = DataValidation(type="list", formula1='"high,medium,low"', allow_blank=True)
    for validation, target in [
        (score_validation, f"E2:E{len(frame)+1}"),
        (role_validation, f"F2:F{len(frame)+1}"),
        (ai_validation, f"G2:G{len(frame)+1}"),
        (confidence_validation, f"H2:H{len(frame)+1}"),
    ]:
        validation.error = "请从允许值中选择，并保证维度与分数映射一致。"
        validation.errorTitle = "输入不合法"
        validation.prompt = "请按‘填写说明’和‘编码规范’工作表填写。"
        validation.promptTitle = "v2.1 人工盲审"
        validation.showErrorMessage = True
        validation.showInputMessage = True
        sheet.add_data_validation(validation)
        validation.add(target)
    sheet.conditional_formatting.add(
        f"A2:J{len(frame)+1}", FormulaRule(formula=["COUNTA($E2:$I2)>0"], fill=PatternFill("solid", fgColor=light_blue))
    )
    sheet.conditional_formatting.add(
        f"E2:I{len(frame)+1}", FormulaRule(formula=['AND($E2<>"",OR($F2="",$G2="",$H2="",AND($E2>0,$I2="")))'], fill=PatternFill("solid", fgColor=pale_red))
    )

    instructions = workbook.create_sheet("填写说明")
    instructions.sheet_view.showGridLines = False
    title = instructions["A1"]
    title.value = f"LLM 文本编码 v2.1 人工盲审（{'开发集' if split == 'development' else '锁定留出集'}）"
    title.font = Font(name="等线", size=16, bold=True, color=navy)
    instructions.merge_cells("A1:F1")
    instruction_rows = [
        ("盲审原则", "只依据‘岗位／岗位描述／岗位标签’判断；不得查找公司、年份或任何模型结果。"),
        ("填写顺序", "先判断 technology_role，再判断 strict_ai；human_score 必须由两者确定，不凭直觉单独打分。"),
        ("确定映射", "none→0；auxiliary→1；core 且 strict_ai=false→2；core 且 strict_ai=true→3。"),
        ("正分证据", "1–3 分必须在 human_evidence 中复制一段原文连续短语；不要改写、拼接或总结。"),
        ("置信度", "high=至少两处一致明确证据且边界清晰；medium=一处明确证据或轻微边界判断；low=信息不足、矛盾或相邻等级均可解释。"),
        ("备注", "human_note 可写边界理由，尤其说明为何不是相邻等级；0 分可留空证据但其他必填字段仍需填写。"),
        ("交付顺序", "先完成并返回开发集。锁定留出集只能在 v2.1 提示词冻结后提交和评测，不能用于调参。"),
        ("禁止操作", "不要新增、删除或重排行，不要修改 review_id 和原文列，不要在两个文件之间复制模型判断。"),
    ]
    instructions.append(["项目", "说明"])
    for label, explanation in instruction_rows:
        instructions.append([label, explanation])
    for row in instructions.iter_rows(min_row=2, max_row=instructions.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.font = Font(name="等线", size=11, bold=cell.column == 1)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
        row[0].fill = PatternFill("solid", fgColor=light_blue)
    instructions.column_dimensions["A"].width = 18
    instructions.column_dimensions["B"].width = 100
    for index in range(2, instructions.max_row + 1):
        instructions.row_dimensions[index].height = 38

    codebook = workbook.create_sheet("编码规范")
    codebook.sheet_view.showGridLines = False
    codebook.append(["分数", "technology_role", "strict_ai", "操作定义", "典型边界"])
    codebook_rows = [
        (0, "none", "false", "没有明示数字技术要求。", "不得依据行业、公司产品或岗位惯例推测。"),
        (1, "auxiliary", "false", "数字工具辅助非数字主业。", "CAD／ERP／Office 常规使用；只协调 AI 项目；普通报表。"),
        (2, "core", "false", "软件、数据、系统、数字硬件或非 AI 算法是主要产出。", "技术售前方案、数字测试、运维安全、核心数据分析、传统算法。"),
        (3, "core", "true", "明确承担 AI／ML 模型研发、训练、推理、评估或部署。", "AI 词仅为公司／产品背景或项目协调时不能判 3。"),
    ]
    for row in codebook_rows:
        codebook.append(row)
    for cell in codebook[1]:
        cell.font = Font(name="等线", size=11, bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in codebook.iter_rows(min_row=2, max_row=5, min_col=1, max_col=5):
        for cell in row:
            cell.font = Font(name="等线", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    for column, width in {"A": 10, "B": 20, "C": 12, "D": 55, "E": 65}.items():
        codebook.column_dimensions[column].width = width
    for index in range(2, 6):
        codebook.row_dimensions[index].height = 48

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(path)


def validate_blind_frame(frame: pd.DataFrame) -> None:
    if list(frame.columns) != REVIEW_COLUMNS:
        raise ValueError(f"blind review columns must be exactly {REVIEW_COLUMNS}")
    forbidden = {"canonical_id", "公司名称", "关联公司名称", "year", "发布时间", "score", "confidence", "selection_reason"}
    leaked = forbidden.intersection(frame.columns)
    if leaked:
        raise ValueError(f"blind review leakage: {sorted(leaked)}")
    if len(frame) != 60 or frame["review_id"].duplicated().any():
        raise ValueError("each blind split must contain 60 unique review IDs")
    if frame[REVIEW_COLUMNS[4:]].ne("").any().any():
        raise ValueError("new blind review input fields must be empty")


def _parse_bool(value: object) -> bool:
    normalized = str(value).strip().casefold()
    if normalized not in {"true", "false"}:
        raise ValueError(f"strict_ai must be true or false, got {value!r}")
    return normalized == "true"


def validate_human_labels(frame: pd.DataFrame) -> pd.DataFrame:
    """Validate a completed single-reviewer reference file and normalize types."""
    missing = [column for column in REVIEW_COLUMNS if column not in frame.columns]
    if missing:
        raise ValueError(f"human review missing columns: {missing}")
    validated = frame[REVIEW_COLUMNS].fillna("").astype(object).copy()
    errors: list[str] = []
    for row_index, row in validated.iterrows():
        review_id = str(row["review_id"])
        try:
            score = int(str(row["human_score"]).strip())
            role = str(row["technology_role"]).strip().casefold()
            strict_ai = _parse_bool(row["strict_ai"])
            confidence = str(row["human_confidence"]).strip().casefold()
            evidence = str(row["human_evidence"]).strip()
            if confidence not in CONFIDENCE_VALUES:
                raise ValueError(f"invalid human_confidence: {confidence}")
            expected_score = score_from_dimensions(role, strict_ai)
            if score != expected_score:
                raise ValueError(f"score {score} conflicts with dimensions; expected {expected_score}")
            source_fields = [str(row["岗位"]), str(row["岗位描述"]), str(row["岗位标签"])]
            if score > 0 and not evidence:
                raise ValueError("positive human_score requires evidence")
            if evidence and not any(evidence in field for field in source_fields):
                raise ValueError("human_evidence is not a contiguous source substring")
            validated.at[row_index, "human_score"] = score
            validated.at[row_index, "technology_role"] = role
            validated.at[row_index, "strict_ai"] = strict_ai
            validated.at[row_index, "human_confidence"] = confidence
        except (TypeError, ValueError) as exc:
            errors.append(f"{review_id}: {exc}")
    if errors:
        raise ValueError("invalid human labels:\n" + "\n".join(errors))
    return validated


def classification_metrics(reference: pd.Series, predicted: pd.Series) -> dict:
    """Return reproducible 0–3, threshold and strict-AI evaluation metrics."""
    actual = pd.Series(reference, dtype=int).reset_index(drop=True)
    model = pd.Series(predicted, dtype=int).reset_index(drop=True)
    if len(actual) != len(model) or not len(actual):
        raise ValueError("reference and prediction must have equal nonzero length")
    confusion = [[int(((actual == truth) & (model == guess)).sum()) for guess in range(4)] for truth in range(4)]
    per_class = {}
    for score in range(4):
        true_positive = int(((actual == score) & (model == score)).sum())
        predicted_positive = int((model == score).sum())
        actual_positive = int((actual == score).sum())
        precision = true_positive / predicted_positive if predicted_positive else None
        recall = true_positive / actual_positive if actual_positive else None
        f1 = (2 * precision * recall / (precision + recall)) if precision is not None and recall is not None and precision + recall else None
        per_class[str(score)] = {
            "support": actual_positive,
            "precision": precision,
            "recall": recall,
            "f1": f1,
        }
    strict = per_class["3"]
    return {
        "sample_size": len(actual),
        "exact_agreement": float(actual.eq(model).mean()),
        "within_one_agreement": float(actual.sub(model).abs().le(1).mean()),
        "binary_agreement_score_ge_2": float(actual.ge(2).eq(model.ge(2)).mean()),
        "quadratic_weighted_kappa": quadratic_weighted_kappa(actual.tolist(), model.tolist()),
        "confusion_matrix_rows_reference_columns_model": confusion,
        "per_class": per_class,
        "strict_ai": {"positives": strict["support"], "precision": strict["precision"], "recall": strict["recall"], "f1": strict["f1"]},
    }


def prepare_human_evaluation(
    *,
    output_dir: Path = Path("artifacts/evals/llm_v2_1"),
    seed: int = DEFAULT_REVIEW_SEED,
) -> dict:
    ads = pd.read_csv("data/processed/cleaned_ads.csv", dtype=str, keep_default_na=False)
    labels = pd.read_csv("outputs/ai_scores.csv", dtype=str, keep_default_na=False)
    comparison = pd.read_csv("artifacts/review/v1_v2_label_comparison.csv", dtype=str, keep_default_na=False)
    reliability = pd.read_csv("artifacts/review/reliability_sample.csv", dtype=str, keep_default_na=False)
    features = build_sampling_frame(ads, labels, comparison, reliability)
    selected = select_human_reference_samples(features, seed)
    output_dir.mkdir(parents=True, exist_ok=True)
    private_dir = output_dir / "private"
    private_dir.mkdir(parents=True, exist_ok=True)
    private_columns = [
        "review_id", "canonical_id", "split", "score", "confidence", "rule_score", "risk_points",
        "score_changed", "crossed_main_threshold", "audit_disagreement", "rule_model_threshold_conflict",
        "low_confidence", "ai_algorithm_candidate", "selection_reason",
    ]
    selected[private_columns].to_csv(private_dir / "review_id_map.csv", index=False, encoding="utf-8-sig")

    split_files: dict[str, dict[str, str]] = {}
    for split, stem in [("development", "blind_development"), ("holdout", "blind_holdout")]:
        blind = blind_review_frame(selected, split, seed)
        validate_blind_frame(blind)
        csv_path = output_dir / f"{stem}.csv"
        xlsx_path = output_dir / f"{stem}.xlsx"
        blind.to_csv(csv_path, index=False, encoding="utf-8-sig")
        write_review_workbook(blind, xlsx_path, split=split)  # type: ignore[arg-type]
        split_files[split] = {"csv": str(csv_path), "xlsx": str(xlsx_path)}

    overlap = set(selected.loc[selected["split"].eq("development"), "canonical_id"]) & set(
        selected.loc[selected["split"].eq("holdout"), "canonical_id"]
    )
    audit = {
        "rubric_version": RUBRIC_VERSION,
        "seed": seed,
        "created_from": {
            "ads": "data/processed/cleaned_ads.csv",
            "labels": "outputs/ai_scores.csv (v2.0.0)",
            "version_comparison": "artifacts/review/v1_v2_label_comparison.csv",
            "same_model_retest": "artifacts/review/reliability_sample.csv",
        },
        "development_count": int(selected["split"].eq("development").sum()),
        "holdout_count": int(selected["split"].eq("holdout").sum()),
        "canonical_id_overlap": len(overlap),
        "blind_columns": REVIEW_COLUMNS,
        "private_mapping_committed": False,
        "score_distribution_by_split": {
            split: selected.loc[selected["split"].eq(split), "score"].value_counts().sort_index().to_dict()
            for split in ["development", "holdout"]
        },
        "files": split_files,
    }
    (output_dir / "selection_audit.json").write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    return audit


def load_boundary_cases(path: Path = Path("config/llm_v2_1_boundary_cases.yaml")) -> list[dict]:
    payload = yaml.safe_load(path.read_text(encoding="utf-8"))
    return list(payload["cases"])
