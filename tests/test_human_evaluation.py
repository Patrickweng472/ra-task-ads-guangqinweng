from pathlib import Path

import pandas as pd
import pytest
import yaml
from openpyxl import load_workbook

from ra_task.human_evaluation import (
    REVIEW_COLUMNS,
    blind_review_frame,
    build_sampling_frame,
    classification_metrics,
    load_boundary_cases,
    normalize_completed_review,
    prepare_human_evaluation,
    score_from_dimensions,
    select_human_reference_samples,
    validate_blind_frame,
    validate_human_labels,
    write_review_workbook,
)


def test_normalize_completed_review_recovers_contiguous_evidence_without_overwriting_source() -> None:
    frame = _completed_review(
        human_evidence=(
            "原文明示‘使用SQL完成经营数据分析’，这说明数据分析是岗位主要产出，"
            "移除数据能力后岗位无法成立，因此为2分而不是1分。"
        ),
        human_note="",
    )

    normalized = normalize_completed_review(frame)

    assert normalized.loc[0, "human_evidence"] == "使用SQL完成经营数据分析"
    assert "因此为2分" in normalized.loc[0, "human_note"]
    validate_human_labels(normalized)


def test_normalize_completed_review_clears_zero_score_evidence_and_preserves_reason_in_note() -> None:
    frame = _completed_review(
        human_score="0",
        technology_role="none",
        strict_ai="false",
        human_evidence="岗位只负责客户接待，因此没有明示数字技术。",
        human_note="",
        岗位="前台",
        岗位描述="负责客户接待",
        岗位标签="行政",
    )

    normalized = normalize_completed_review(frame)

    assert normalized.loc[0, "human_evidence"] == ""
    assert "没有明示数字技术" in normalized.loc[0, "human_note"]


def _real_selection() -> pd.DataFrame:
    ads = pd.read_csv("data/processed/cleaned_ads.csv", dtype=str, keep_default_na=False)
    labels = pd.read_csv("outputs/ai_scores.csv", dtype=str, keep_default_na=False)
    comparison = pd.read_csv("artifacts/review/v1_v2_label_comparison.csv", dtype=str, keep_default_na=False)
    reliability = pd.read_csv("artifacts/review/reliability_sample.csv", dtype=str, keep_default_na=False)
    features = build_sampling_frame(ads, labels, comparison, reliability)
    return select_human_reference_samples(features, seed=20260629)


@pytest.mark.parametrize(
    ("role", "strict_ai", "score"),
    [("none", False, 0), ("auxiliary", False, 1), ("core", False, 2), ("core", True, 3)],
)
def test_score_is_deterministically_derived_from_dimensions(role: str, strict_ai: bool, score: int) -> None:
    assert score_from_dimensions(role, strict_ai) == score


def test_invalid_strict_ai_dimension_combination_is_rejected() -> None:
    with pytest.raises(ValueError, match="requires"):
        score_from_dimensions("auxiliary", True)


def test_v2_1_boundary_suite_has_at_least_thirty_consistent_cases() -> None:
    cases = load_boundary_cases()
    assert len(cases) >= 30
    assert len({case["id"] for case in cases}) == len(cases)
    required_topics = {
        "technical_presales", "ai_coordination", "engineering_tools", "system_operations",
        "digital_testing", "data_work", "traditional_algorithms", "strict_ai",
    }
    assert required_topics.issubset({case["topic"] for case in cases})
    for case in cases:
        assert case["score"] == score_from_dimensions(case["technology_role"], bool(case["strict_ai"]))


def test_v2_1_rubric_freezes_critical_boundary_rules() -> None:
    rubric = yaml.safe_load(Path("config/ai_rubric_v2_1.yaml").read_text(encoding="utf-8"))
    assert rubric["version"] == "2.1.0"
    assert set(rubric["boundary_rules"]) == {
        "technical_presales", "testing", "engineering_tools", "ai_coordination", "data_work", "traditional_algorithms"
    }
    assert "移除" in rubric["core_duty_test"]
    assert "AI" in rubric["boundary_rules"]["ai_coordination"]


def test_real_reference_selection_is_disjoint_reproducible_and_quota_balanced() -> None:
    first = _real_selection()
    second = _real_selection()
    pd.testing.assert_frame_equal(first.reset_index(drop=True), second.reset_index(drop=True))
    assert len(first) == 120
    assert not first["canonical_id"].duplicated().any()
    assert first.groupby("split").size().to_dict() == {"development": 60, "holdout": 60}
    assert not set(first.loc[first["split"].eq("development"), "canonical_id"]) & set(
        first.loc[first["split"].eq("holdout"), "canonical_id"]
    )
    distributions = first.groupby(["split", "score"]).size().to_dict()
    assert distributions == {
        ("development", 0): 15, ("development", 1): 20, ("development", 2): 24, ("development", 3): 1,
        ("holdout", 0): 18, ("holdout", 1): 20, ("holdout", 2): 21, ("holdout", 3): 1,
    }
    assert int(first.loc[first["split"].eq("development"), "ai_algorithm_candidate"].sum()) >= 6
    assert int(first.loc[first["split"].eq("holdout"), "ai_algorithm_candidate"].sum()) >= 4


def test_blind_frames_have_exact_columns_and_no_sensitive_metadata() -> None:
    selected = _real_selection()
    development = blind_review_frame(selected, "development", seed=20260629)
    holdout = blind_review_frame(selected, "holdout", seed=20260629)
    validate_blind_frame(development)
    validate_blind_frame(holdout)
    assert list(development.columns) == REVIEW_COLUMNS
    forbidden = {"canonical_id", "公司名称", "year", "score", "confidence", "selection_reason"}
    assert not forbidden.intersection(development.columns)
    assert not set(development["review_id"]) & set(holdout["review_id"])
    assert development[REVIEW_COLUMNS[4:]].eq("").all().all()


def test_review_workbook_is_blind_styled_and_contains_no_formulas(tmp_path: Path) -> None:
    blind = blind_review_frame(_real_selection(), "development", seed=20260629)
    path = tmp_path / "review.xlsx"
    write_review_workbook(blind, path, split="development")
    workbook = load_workbook(path, data_only=False)
    assert workbook.sheetnames == ["待审核", "填写说明", "编码规范"]
    sheet = workbook["待审核"]
    assert sheet.freeze_panes == "E2"
    assert sheet.auto_filter.ref == "A1:J61"
    assert [cell.value for cell in sheet[1]] == REVIEW_COLUMNS
    assert all(cell.font.name == "等线" for worksheet in workbook.worksheets for row in worksheet.iter_rows() for cell in row if cell.value is not None)
    assert not [cell.coordinate for worksheet in workbook.worksheets for row in worksheet.iter_rows() for cell in row if cell.data_type == "f"]
    all_text = "\n".join(str(cell.value or "") for worksheet in workbook.worksheets for row in worksheet.iter_rows() for cell in row)
    for forbidden in ["公司名称", "发布时间", "selection_reason", "primary_score", "canonical_id"]:
        assert forbidden not in all_text


def test_prepare_human_evaluation_writes_complete_public_bundle_and_private_map(tmp_path: Path) -> None:
    output = tmp_path / "eval"
    audit = prepare_human_evaluation(output_dir=output, seed=20260629)
    assert audit["development_count"] == 60
    assert audit["holdout_count"] == 60
    assert audit["canonical_id_overlap"] == 0
    expected = {
        "blind_development.csv", "blind_development.xlsx", "blind_holdout.csv", "blind_holdout.xlsx",
        "selection_audit.json", "private/review_id_map.csv",
    }
    assert expected.issubset({str(path.relative_to(output)) for path in output.rglob("*") if path.is_file()})
    for split in ["development", "holdout"]:
        blind = pd.read_csv(output / f"blind_{split}.csv", dtype=str, keep_default_na=False)
        validate_blind_frame(blind)


def _completed_review(**changes: object) -> pd.DataFrame:
    row = {
        "review_id": "DEV-ABC",
        "岗位": "数据分析师",
        "岗位描述": "使用SQL完成经营数据分析",
        "岗位标签": "数据",
        "human_score": "2",
        "technology_role": "core",
        "strict_ai": "false",
        "human_confidence": "high",
        "human_evidence": "使用SQL完成经营数据分析",
        "human_note": "数据分析是主要产出",
    }
    row.update(changes)
    return pd.DataFrame([row], columns=REVIEW_COLUMNS)


def test_completed_human_review_validates_and_normalizes() -> None:
    result = validate_human_labels(_completed_review())
    assert result.loc[0, "human_score"] == 2
    assert result.loc[0, "strict_ai"] is False or result.loc[0, "strict_ai"] == False  # noqa: E712


@pytest.mark.parametrize(
    ("changes", "message"),
    [
        ({"human_score": "3"}, "conflicts"),
        ({"strict_ai": "true"}, "conflicts"),
        ({"human_evidence": "改写后的证据"}, "contiguous"),
        ({"human_evidence": ""}, "requires evidence"),
        ({"human_confidence": "certain"}, "invalid human_confidence"),
    ],
)
def test_invalid_human_reference_rows_are_rejected(changes: dict, message: str) -> None:
    with pytest.raises(ValueError, match=message):
        validate_human_labels(_completed_review(**changes))


def test_evaluation_metrics_include_confusion_per_class_threshold_and_strict_ai() -> None:
    result = classification_metrics(pd.Series([0, 1, 2, 3]), pd.Series([0, 2, 2, 3]))
    assert result["exact_agreement"] == pytest.approx(0.75)
    assert result["binary_agreement_score_ge_2"] == pytest.approx(0.75)
    assert result["confusion_matrix_rows_reference_columns_model"][1][2] == 1
    assert result["per_class"]["2"]["precision"] == pytest.approx(0.5)
    assert result["strict_ai"]["f1"] == pytest.approx(1.0)
